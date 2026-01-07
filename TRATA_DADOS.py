import pandas as pd
pd.set_option("future.no_silent_downcasting", True)
from pathlib import Path
from tqdm import tqdm
import time
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

inicio = time.time()

# ==============================
# CONFIGURAÇÕES
# ==============================
BASE_DIR = Path(__file__).resolve().parent

# ATENÇÃO: Os arquivos em .csv extaídos do Aspect devem estar nesta pasta:
PASTA_BASE = BASE_DIR / "BASES_RAW"

# Aqui é onde serão salvos os arquivos, da base tratada e o relatório final.
PASTA_SAIDA = BASE_DIR / "BASE_TRATADA"

PASTA_BASE.mkdir(exist_ok=True)
PASTA_SAIDA.mkdir(exist_ok=True)

ARQUIVO_TRATADO = PASTA_SAIDA / "base_tratada.csv"
ARQUIVO_CONSOLIDADO = PASTA_SAIDA / "relatorio_completo.xlsx"

# ==============================
# COLUNAS
# ==============================
COLUNAS = [
    "CallStartDt",
    "SeqNum",
    "CallId",
    "DetectionDt",
    "AnswerDt",
    "WrapEndDt",
    "CallInsertDt",
    "CallEndDt",
    "TimePhoneStartingRinging",
    "DialedNum",
    "Disp_c",
    "Disposition_Desc",
    "ResourceGroupDesc"
]

# ==============================
# 1. LEITURA E UNIFICAÇÃO
# ==============================
arquivos = sorted(PASTA_BASE.glob("*.csv"))
dfs = []

print(f"Iniciando leitura de {len(arquivos)} arquivos...")

for arquivo in tqdm(arquivos, desc="📂 Lendo arquivos", unit="arquivo"):
    try:
        df = pd.read_csv(
            arquivo,
            sep=";",
            header=None,
            names=COLUNAS,
            encoding="utf-8",
            dtype=str,
            low_memory=False,
            on_bad_lines="skip"
        )

        df = df.dropna(how="all")

        # normaliza nulos logo na origem também
        df = df.replace({"NULL": pd.NA, "null": pd.NA, "": pd.NA, " ": pd.NA})
        df = df.infer_objects(copy=False)

        # filtra linhas sem CallStartDt
        df = df[df["CallStartDt"].notna()]

        dfs.append(df)
        print(f"  ✓ {arquivo.name}: {len(df)} registros válidos")

    except Exception as e:
        print(f"  ✗ ERRO ao ler {arquivo.name}: {e}")
        continue

if not dfs:
    raise ValueError("❌ Nenhum arquivo foi carregado com sucesso!")

cdr = pd.concat(dfs, ignore_index=True)
print(f"\n✅ Total de registros carregados: {len(cdr)}")

# ==============================
# 2. ANÁLISE DE TIPAGEM (ANTES DA CONVERSÃO)
# ==============================
log_tipagem = []
amostra_size = min(10000, len(cdr))
amostra = cdr.sample(n=amostra_size, random_state=42)

for coluna in COLUNAS:
    valores_unicos = amostra[coluna].nunique(dropna=True)
    valores_nulos = amostra[coluna].isna().sum()
    valores_preenchidos = amostra_size - valores_nulos

    tipo_detectado = "string"
    valores_validos = amostra[coluna].dropna()

    if len(valores_validos) > 0:
        try:
            pd.to_numeric(valores_validos, errors="raise")
            tipo_detectado = "numeric"
        except:
            try:
                with pd.option_context("mode.chained_assignment", None):
                    converted = pd.to_datetime(valores_validos, errors="coerce", format="mixed")
                if converted.notna().sum() / len(valores_validos) > 0.8:
                    tipo_detectado = "datetime"
            except:
                pass

    exemplos = valores_validos.head(3).tolist()
    exemplos_str = " | ".join([str(x)[:30] for x in exemplos])

    log_tipagem.append({
        "coluna": coluna,
        "tipo_detectado": tipo_detectado,
        "valores_unicos": valores_unicos,
        "valores_nulos": valores_nulos,
        "valores_preenchidos": valores_preenchidos,
        "perc_preenchimento": round(valores_preenchidos / amostra_size * 100, 2),
        "exemplos": exemplos_str
    })

df_tipagem = pd.DataFrame(log_tipagem)

# ==============================
# 3. NORMALIZA NULOS (geral)
# ==============================
print("\n🔄 Normalizando valores NULL...")
cdr = cdr.replace({"NULL": pd.NA, "null": pd.NA, "": pd.NA, " ": pd.NA})
cdr = cdr.infer_objects(copy=False)

# ==============================
# 4. TIPAGEM DE DADOS
# ==============================
colunas_datetime = [
    "CallStartDt",
    "DetectionDt",
    "AnswerDt",
    "WrapEndDt",
    "CallInsertDt",
    "CallEndDt",
    "TimePhoneStartingRinging"
]

# Conversão de datas
for col in colunas_datetime:
    print(f"  → {col}...", end=" ")
    cdr[col] = pd.to_datetime(cdr[col], errors="coerce")
    nulls = cdr[col].isna().sum()
    validos = len(cdr) - nulls
    print(f"{validos:,} válidos, {nulls:,} nulos")

# Numéricos
cdr["SeqNum"] = pd.to_numeric(cdr["SeqNum"], errors="coerce")
cdr["CallId"] = pd.to_numeric(cdr["CallId"], errors="coerce")

# ==============================
# 5. VERIFICAÇÃO DE QUALIDADE PRÉ-CÁLCULO
# ==============================
print("\n🔍 Verificando dados antes dos cálculos...")
print(f"  CallStartDt válidos: {cdr['CallStartDt'].notna().sum():,}")
print(f"  AnswerDt válidos: {cdr['AnswerDt'].notna().sum():,}")
print(f"  WrapEndDt válidos: {cdr['WrapEndDt'].notna().sum():,}")
print(f"  TimePhoneStartingRinging válidos: {cdr['TimePhoneStartingRinging'].notna().sum():,}")

# ==============================
# 6. MÉTRICAS DE TEMPO
# ==============================
cdr["ring_time_sec"] = (cdr["AnswerDt"] - cdr["TimePhoneStartingRinging"]).dt.total_seconds()
cdr["talk_time_sec"] = (cdr["WrapEndDt"] - cdr["AnswerDt"]).dt.total_seconds()
cdr["call_duration_sec"] = (cdr["CallEndDt"] - cdr["CallStartDt"]).dt.total_seconds()
cdr["wrap_time_sec"] = (cdr["WrapEndDt"] - cdr["CallEndDt"]).dt.total_seconds()

# invalida negativos
for col in ["ring_time_sec", "talk_time_sec", "call_duration_sec", "wrap_time_sec"]:
    cdr.loc[cdr[col] < 0, col] = pd.NA

# ==============================
# 7. FLAGS DE NEGÓCIO (SLA 0/1 garantido)
# ==============================
cdr["atendida"] = cdr["AnswerDt"].notna().astype("int8")

# ring válido: existe, não é NA e >=0
ring_valido = cdr["ring_time_sec"].notna()

# SLA: somente se atendida e ring_valido e dentro do limite
cdr["sla_15s"] = ((cdr["atendida"] == 1) & ring_valido & (cdr["ring_time_sec"] <= 15)).astype("int8")
cdr["sla_30s"] = ((cdr["atendida"] == 1) & ring_valido & (cdr["ring_time_sec"] <= 30)).astype("int8")

# SLA nunca pode ser 1 se não atendida
cdr.loc[cdr["atendida"] == 0, ["sla_15s", "sla_30s"]] = 0

cdr["hora"] = cdr["CallStartDt"].dt.hour
cdr["data"] = cdr["CallStartDt"].dt.date

# ==============================
# 7.1 NORMALIZA GRUPO DE RECURSOS
# ==============================
cdr["ResourceGroupDesc"] = cdr["ResourceGroupDesc"].fillna("SEM_GRUPO").astype(str).str.strip()
cdr.loc[cdr["ResourceGroupDesc"].eq(""), "ResourceGroupDesc"] = "SEM_GRUPO"

# ==============================
# 8 CHAVE ÚNICA (evita 'nan_nan')
# ==============================
print("\n🔑 Criando chave única...")

cdr["CallId_int"] = cdr["CallId"].astype("Int64")
cdr["SeqNum_int"] = cdr["SeqNum"].astype("Int64")

cdr["chave_unica"] = (
    cdr["CallId_int"].astype(str).replace("<NA>", "SEM_CALLID")
    + "_"
    + cdr["SeqNum_int"].astype(str).replace("<NA>", "SEM_SEQNUM")
)

# remove colunas auxiliares
cdr.drop(columns=["CallId_int", "SeqNum_int"], inplace=True)

cols = cdr.columns.tolist()
cols = ["chave_unica"] + [col for col in cols if col != "chave_unica"]
cdr = cdr[cols]

print(f"   ✓ Chave única criada: {cdr['chave_unica'].nunique():,} registros únicos")

# ==============================
# 9 DETECÇÃO DE ANOMALIAS
# ==============================
print("\n🚨 Detectando anomalias (hora e grupo)...")

def zscore(series: pd.Series) -> pd.Series:
    s = pd.to_numeric(series, errors="coerce")
    mean = s.mean()
    std = s.std(ddof=0)
    if pd.isna(std) or std == 0:
        return pd.Series([0] * len(s), index=s.index)
    return (s - mean) / std

# ---------- (A) Anomalias por HORA ----------
agg_hora = (
    cdr.groupby("hora", dropna=False)
       .agg(
           total_chamadas=("chave_unica", "count"),
           atendidas=("atendida", "sum"),
           sla_15=("sla_15s", "sum"),
           sla_30=("sla_30s", "sum"),
           ring_medio=("ring_time_sec", "mean"),
           talk_medio=("talk_time_sec", "mean"),
       )
       .reset_index()
)

agg_hora["taxa_atendimento"] = np.where(
    agg_hora["total_chamadas"] > 0,
    agg_hora["atendidas"] / agg_hora["total_chamadas"],
    np.nan
)

agg_hora["sla15_rate"] = np.where(
    agg_hora["atendidas"] > 0,
    agg_hora["sla_15"] / agg_hora["atendidas"],
    np.nan
)

agg_hora["sla30_rate"] = np.where(
    agg_hora["atendidas"] > 0,
    agg_hora["sla_30"] / agg_hora["atendidas"],
    np.nan
)

agg_hora["z_volume"] = zscore(agg_hora["total_chamadas"])
agg_hora["z_taxa_atendimento"] = zscore(agg_hora["taxa_atendimento"])

VOLUME_K = 1.5
TAXA_Z = -2.0

if agg_hora["total_chamadas"].count() < 8:
    print("⚠️ Poucos pontos horários para z-score confiável. Flags de anomalia por HORA serão False.")
    agg_hora["anomalia_volume"] = False
    agg_hora["anomalia_taxa_atendimento"] = False
else:
    agg_hora["anomalia_volume"] = agg_hora["z_volume"].abs() >= VOLUME_K
    agg_hora["anomalia_taxa_atendimento"] = agg_hora["z_taxa_atendimento"] <= TAXA_Z

agg_hora["flag_anomalia"] = agg_hora["anomalia_volume"] | agg_hora["anomalia_taxa_atendimento"]

agg_hora["motivo_anomalia"] = ""
agg_hora.loc[agg_hora["anomalia_volume"], "motivo_anomalia"] += "VOLUME_FORA_PADRAO; "
agg_hora.loc[agg_hora["anomalia_taxa_atendimento"], "motivo_anomalia"] += "TAXA_ATENDIMENTO_BAIXA; "
agg_hora["motivo_anomalia"] = agg_hora["motivo_anomalia"].str.strip()

# ---------- (B) Anomalias por GRUPO ----------
agg_grupo = (
    cdr.groupby("ResourceGroupDesc", dropna=False)
       .agg(
           total_chamadas=("chave_unica", "count"),
           atendidas=("atendida", "sum"),
           sla_15=("sla_15s", "sum"),
           sla_30=("sla_30s", "sum"),
           ring_medio=("ring_time_sec", "mean"),
           talk_medio=("talk_time_sec", "mean"),
       )
       .reset_index()
)

agg_grupo["taxa_atendimento"] = np.where(
    agg_grupo["total_chamadas"] > 0,
    agg_grupo["atendidas"] / agg_grupo["total_chamadas"],
    np.nan
)

agg_grupo["sla15_rate"] = np.where(
    agg_grupo["atendidas"] > 0,
    agg_grupo["sla_15"] / agg_grupo["atendidas"],
    np.nan
)

agg_grupo["sla30_rate"] = np.where(
    agg_grupo["atendidas"] > 0,
    agg_grupo["sla_30"] / agg_grupo["atendidas"],
    np.nan
)

agg_grupo["z_volume"] = zscore(agg_grupo["total_chamadas"])
agg_grupo["z_taxa_atendimento"] = zscore(agg_grupo["taxa_atendimento"])

agg_grupo["anomalia_volume"] = agg_grupo["z_volume"].abs() >= VOLUME_K
agg_grupo["anomalia_taxa_atendimento"] = agg_grupo["z_taxa_atendimento"] <= TAXA_Z
agg_grupo["flag_anomalia"] = agg_grupo["anomalia_volume"] | agg_grupo["anomalia_taxa_atendimento"]

agg_grupo["motivo_anomalia"] = ""
agg_grupo.loc[agg_grupo["anomalia_volume"], "motivo_anomalia"] += "VOLUME_FORA_PADRAO; "
agg_grupo.loc[agg_grupo["anomalia_taxa_atendimento"], "motivo_anomalia"] += "TAXA_ATENDIMENTO_BAIXA; "
agg_grupo["motivo_anomalia"] = agg_grupo["motivo_anomalia"].str.strip()

# ---------- (C) Anomalias por HORA×GRUPO ----------
agg_hora_grupo = (
    cdr.groupby(["hora", "ResourceGroupDesc"], dropna=False)
       .agg(total_chamadas=("chave_unica", "count"), atendidas=("atendida", "sum"))
       .reset_index()
)

agg_hora_grupo["taxa_atendimento"] = np.where(
    agg_hora_grupo["total_chamadas"] > 0,
    agg_hora_grupo["atendidas"] / agg_hora_grupo["total_chamadas"],
    np.nan
)

MIN_CHAMADAS_PARA_ANALISE = 30
base_hg = agg_hora_grupo[agg_hora_grupo["total_chamadas"] >= MIN_CHAMADAS_PARA_ANALISE].copy()

base_hg["z_taxa_atendimento_no_grupo"] = (
    base_hg.groupby("ResourceGroupDesc")["taxa_atendimento"].transform(lambda s: zscore(s))
)

base_hg["anomalia_taxa_no_grupo"] = base_hg["z_taxa_atendimento_no_grupo"] <= TAXA_Z

print(f"   ✓ Horas com anomalia: {int(agg_hora['flag_anomalia'].sum())}")
print(f"   ✓ Grupos com anomalia: {int(agg_grupo['flag_anomalia'].sum())}")

# ==============================
# 10. DISTRIBUIÇÃO POR DISPOSITION
# ==============================
cdr["Disposition_Desc"] = cdr["Disposition_Desc"].fillna("SEM_DISPOSITION").astype(str).str.strip()
cdr.loc[cdr["Disposition_Desc"].eq(""), "Disposition_Desc"] = "SEM_DISPOSITION"

df_disp = (
    cdr.groupby("Disposition_Desc", dropna=False)
       .size()
       .reset_index(name="total_chamadas")
       .sort_values("total_chamadas", ascending=False)
)

df_disp["perc_total"] = (df_disp["total_chamadas"] / len(cdr) * 100).round(2)

df_disp_hora = (
    cdr.groupby(["hora", "Disposition_Desc"], dropna=False)
       .size()
       .reset_index(name="total_chamadas")
       .sort_values(["hora", "total_chamadas"], ascending=[True, False])
)

total_por_hora = cdr.groupby("hora").size().reset_index(name="total_hora")
df_disp_hora = df_disp_hora.merge(total_por_hora, on="hora", how="left")
df_disp_hora["perc_na_hora"] = (df_disp_hora["total_chamadas"] / df_disp_hora["total_hora"] * 100).round(2)

# ==============================
# 11. QUALIDADE DE DADOS
# ==============================
resumo_dados = []
log_qualidade = []

wrap_medio = cdr["wrap_time_sec"].dropna().mean()

def registrar(categoria, metrica, valor, percentual=None, severidade="INFO"):
    reg = {
        "categoria": categoria,
        "metrica": metrica,
        "valor": int(valor) if pd.notna(valor) else 0,
        "severidade": severidade
    }
    if percentual is not None:
        reg["percentual"] = round(percentual, 2)
    log_qualidade.append(reg)

total = len(cdr)

registrar("VALORES_AUSENTES", "total_registros", total, severidade="INFO")
registrar("VALORES_AUSENTES", "callid_nulo", cdr["CallId"].isna().sum(),
          (cdr["CallId"].isna().sum() / total * 100),
          "CRÍTICO" if cdr["CallId"].isna().sum() > 0 else "OK")
registrar("VALORES_AUSENTES", "callstartdt_nulo", cdr["CallStartDt"].isna().sum(),
          (cdr["CallStartDt"].isna().sum() / total * 100),
          "CRÍTICO" if cdr["CallStartDt"].isna().sum() > 0 else "OK")
registrar("TEMPOS", "wrap_time_nulo", cdr["wrap_time_sec"].isna().sum(),
          (cdr["wrap_time_sec"].isna().sum()/total*100), "ALERTA")

duplicatas_chave_unica = cdr["chave_unica"].duplicated().sum()
registrar("DUPLICIDADE", "duplicatas_chave_unica", duplicatas_chave_unica,
          (duplicatas_chave_unica / total * 100),
          "CRÍTICO" if duplicatas_chave_unica > total * 0.01 else "ALERTA" if duplicatas_chave_unica > 0 else "OK")

df_qualidade = pd.DataFrame(log_qualidade)

# ==============================
# 12. SALVAR CSV TRATADO
# ==============================
print("\n💾 Salvando CSV tratado...")
cdr.to_csv(ARQUIVO_TRATADO, index=False, encoding="utf-8-sig")
print(f"   ✓ CDR tratado salvo: {ARQUIVO_TRATADO.name}")

# ==============================
# 13. RESUMO EXECUTIVO
# ==============================
print("\n📊 Gerando resumo executivo...")

print("\n" + "=" * 50)
print("📈 RESUMO EXECUTIVO (CONSOLE)")
print("=" * 50)

total = len(cdr)
unicos = cdr["chave_unica"].nunique() if "chave_unica" in cdr.columns else 0

atendidas = int(cdr["atendida"].sum()) if "atendida" in cdr.columns else 0
nao_atendidas = total - atendidas
taxa_at = (atendidas / total) if total else 0

print(f"   • Total de chamadas: {total:,}")
print(f"   • Registros únicos (chave_unica): {unicos:,} ({(unicos/total*100 if total else 0):.1f}%)")
print(f"   • Chamadas atendidas: {atendidas:,} ({taxa_at*100:.1f}%)")
print(f"   • Chamadas não atendidas: {nao_atendidas:,} ({(nao_atendidas/total*100 if total else 0):.1f}%)")

if atendidas > 0:
    sla_15 = int(cdr["sla_15s"].sum()) if "sla_15s" in cdr.columns else 0
    sla_30 = int(cdr["sla_30s"].sum()) if "sla_30s" in cdr.columns else 0

    print(f"   • SLA 15s: {sla_15:,} ({(sla_15/atendidas*100):.1f}% das atendidas)")
    print(f"   • SLA 30s: {sla_30:,} ({(sla_30/atendidas*100):.1f}% das atendidas)")

    if "ring_time_sec" in cdr.columns and cdr["ring_time_sec"].notna().any():
        print(f"   • Ring time médio: {cdr['ring_time_sec'].dropna().mean():.1f}s")

    if "talk_time_sec" in cdr.columns and cdr["talk_time_sec"].notna().any():
        print(f"   • Talk time médio: {cdr['talk_time_sec'].dropna().mean():.1f}s")

    if "wrap_time_sec" in cdr.columns and cdr["wrap_time_sec"].notna().any():
        print(f"   • Wrap time médio: {cdr['wrap_time_sec'].dropna().mean():.1f}s")
else:
    print("   ⚠️  ATENÇÃO: Nenhuma chamada atendida detectada!")

if "df_qualidade" in globals() and isinstance(df_qualidade, pd.DataFrame) and "severidade" in df_qualidade.columns:
    criticos = df_qualidade[df_qualidade["severidade"] == "CRÍTICO"]
    alertas = df_qualidade[df_qualidade["severidade"] == "ALERTA"]

    print("\n⚠️  ALERTAS DE QUALIDADE:")
    if len(criticos) > 0:
        print(f"   🔴 {len(criticos)} problemas CRÍTICOS detectados:")
        for _, row in criticos.iterrows():
            print(f"      • {row.get('metrica','-')}: {int(row.get('valor',0)):,} ocorrências")
    else:
        print("   ✅ Nenhum problema crítico detectado")

    if len(alertas) > 0:
        print(f"   🟡 {len(alertas)} alertas detectados:")
        for _, row in alertas.iterrows():
            print(f"      • {row.get('metrica','-')}: {int(row.get('valor',0)):,} ocorrências")
    else:
        print("   ✅ Nenhum alerta detectado")

print("=" * 50)

# Salva no Dataframe
resumo_dados.append({
    "Categoria": "GERAL",
    "Métrica": "Total de chamadas",
    "Valor": len(cdr),
    "Percentual": "100.00%",
    "Status": "✓"
})

chamadas_atendidas = int(cdr["atendida"].sum())
perc_atendidas = (chamadas_atendidas / len(cdr) * 100) if len(cdr) else 0

resumo_dados.append({
    "Categoria": "GERAL",
    "Métrica": "Chamadas atendidas",
    "Valor": chamadas_atendidas,
    "Percentual": f"{perc_atendidas:.1f}%",
    "Status": "✓" if perc_atendidas > 20 else "⚠"
})

resumo_dados.append({
    "Categoria": "TEMPOS",
    "Métrica": "Wrap médio (s)",
    "Valor": round(wrap_medio, 2) if pd.notna(wrap_medio) else None,
    "Percentual": "",
    "Status": "✓"
})

if chamadas_atendidas > 0:
    sla_15 = int(cdr["sla_15s"].sum())
    perc_sla_15 = (sla_15 / chamadas_atendidas * 100)

    resumo_dados.append({
        "Categoria": "SLA",
        "Métrica": "SLA 15s (das atendidas)",
        "Valor": sla_15,
        "Percentual": f"{perc_sla_15:.1f}%",
        "Status": "✓" if perc_sla_15 >= 80 else "⚠"
    })

df_resumo = pd.DataFrame(resumo_dados)

# ==============================
# 14. RESUMO DA ANÁLISE (EXCEL)
# ==============================
def salvar_excel_consolidado_com_fallback(caminho_base: Path):
    try:
        writer = pd.ExcelWriter(caminho_base, engine="openpyxl")
        return writer, caminho_base
    except PermissionError:
        timestamp = datetime.datetime.now().strftime("%H%M%S")
        novo_nome = caminho_base.stem + f"_{timestamp}" + caminho_base.suffix
        novo_caminho = caminho_base.parent / novo_nome
        writer = pd.ExcelWriter(novo_caminho, engine="openpyxl")
        print(f"   ⚠️  Arquivo em uso! Salvando como: {novo_caminho.name}")
        return writer, novo_caminho

def estilizar_botao(cell, fill_color="D9E1F2", font_color="1F4E79"):
    cell.font = Font(bold=True, color=font_color)
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin", color="A6A6A6"),
        right=Side(style="thin", color="A6A6A6"),
        top=Side(style="thin", color="A6A6A6"),
        bottom=Side(style="thin", color="A6A6A6"),
    )

def aplicar_estilo_tabela(ws, header_color="2F5597"):
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")

    zebra1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    zebra2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    align = Alignment(vertical="center", wrap_text=True)

    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    max_row = ws.max_row
    max_col = ws.max_column

    header_row = 1

    for c in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for r in range(header_row + 1, max_row + 1):
        fill = zebra1 if (r % 2 == 0) else zebra2
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.alignment = align
            cell.border = border

            if isinstance(cell.value, str):
                v = cell.value.strip().upper()
                if v in ("CRÍTICO", "CRITICO", "✗"):
                    cell.font = Font(bold=True, color="C00000")
                elif v in ("ALERTA", "⚠"):
                    cell.font = Font(bold=True, color="FF6600")
                elif v in ("OK", "✓"):
                    cell.font = Font(bold=True, color="008000")

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = f"A{header_row + 1}"
    ws.row_dimensions[header_row].height = 22

    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        max_len = 0
        for r in range(1, max_row + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[letter].width = min(max_len + 2, 55)

    rate_cols = []
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        name = str(v).strip().lower()
        if name.endswith("_rate") or name in ("taxa_atendimento", "sla15_rate", "sla30_rate"):
            rate_cols.append(c)

    for c in rate_cols:
        for r in range(header_row + 1, max_row + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

def criar_dashboard(ws, nomes_abas):
    ws["A1"] = "Dashboard - Navegação"
    ws["A1"].font = Font(size=16, bold=True, color="000000")
    ws["A1"].fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 28

    ws["A3"] = "Clique para abrir:"
    ws["A3"].font = Font(bold=True, color="1F4E79")

    row = 5
    for aba in nomes_abas:
        cell = ws[f"A{row}"]
        cell.value = f"📄 {aba}"
        cell.hyperlink = f"#'{aba}'!A1"
        cell.style = "Hyperlink"
        estilizar_botao(cell, fill_color="E2EFDA", font_color="1F4E79")
        ws.merge_cells(f"A{row}:C{row}")
        ws.row_dimensions[row].height = 22
        row += 2

    ws.column_dimensions["A"].width = 35
    ws.freeze_panes = "A5"

print("\n💾 Gerando Excel consolidado (1 arquivo)...")

writer, caminho_consolidado_final = salvar_excel_consolidado_com_fallback(ARQUIVO_CONSOLIDADO)
with writer:
    df_resumo.to_excel(writer, sheet_name="Resumo Executivo", index=False)
    df_qualidade.to_excel(writer, sheet_name="Qualidade de Dados", index=False)
    df_tipagem.to_excel(writer, sheet_name="Tipagem de Colunas", index=False)
    agg_hora.sort_values("hora").to_excel(writer, sheet_name="Anomalias - Hora", index=False)
    agg_grupo.sort_values("total_chamadas", ascending=False).to_excel(writer, sheet_name="Anomalias - Grupo", index=False)
    base_hg.sort_values(["ResourceGroupDesc", "hora"]).to_excel(writer, sheet_name="Anomalias - Hora×Grupo", index=False)
    df_disp.to_excel(writer, sheet_name="Disposition - Geral", index=False)
    df_disp_hora.to_excel(writer, sheet_name="Disposition - Hora", index=False)

wb = load_workbook(caminho_consolidado_final)

abas_para_link = [
    "Resumo Executivo",
    "Qualidade de Dados",
    "Tipagem de Colunas",
    "Anomalias - Hora",
    "Anomalias - Grupo",
    "Anomalias - Hora×Grupo",
    "Disposition - Geral",
    "Disposition - Hora",
]

ws_dash = wb.create_sheet("Dashboard", 0)
criar_dashboard(ws_dash, abas_para_link)

for nome_aba, cor in [
    ("Resumo Executivo", "1F4E79"),
    ("Qualidade de Dados", "C00000"),
    ("Tipagem de Colunas", "70AD47"),
    ("Anomalias - Hora", "305496"),
    ("Anomalias - Grupo", "548235"),
    ("Anomalias - Hora×Grupo", "BF8F00"),
    ("Disposition - Geral", "5B9BD5"),
    ("Disposition - Hora", "5B9BD5"),
]:
    ws = wb[nome_aba]
    aplicar_estilo_tabela(ws, cor)

ws_dash.column_dimensions["A"].width = 35
wb.save(caminho_consolidado_final)
print(f"   ✓ Relatório consolidado salvo em: {caminho_consolidado_final.name}")

print("\n" + "=" * 50)
print("✅ PROCESSAMENTO FINALIZADO COM SUCESSO")
print("=" * 50)
print(f"📄 Arquivo tratado (CSV): {ARQUIVO_TRATADO.name}")
print(f"📊 Relatório (Excel): {caminho_consolidado_final.name}")
print("=" * 50)

fim = time.time()
print(f"\n⏱ Tempo total de execução: {round((fim - inicio) / 60, 2)} minutos")
print("\n🎯 Pronto para análise!")
